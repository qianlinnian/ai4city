"""
从 filled_metrics.xlsx 与分析图目录加载场景数据。
按图片文件名（或 stem）前 26 位匹配 B 列唯一行。
"""

from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config import (
    ASSETS_DIR,
    EDGE_MAPS_DIR,
    EXPERIENCE_KEYS,
    FILLED_METRICS_XLSX,
    IMAGE_KEY_LEN,
    SEG_MAPS_DIR,
    SKYLINE_MAPS_DIR,
)

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".JPG", ".JPEG", ".PNG", ".WEBP"}

# Excel 列（1-based）：情景 C–I；形态 J–P（注意 L=人造物、M=天空，与 MORPH_KEYS 顺序不同）
_SCENE_COL_MAP = {
    3: "observation_time",
    4: "observation_weather",
    5: "people_flow",
    6: "space_type",
    7: "sound_type",
    8: "maintenance_status",
    9: "traffic_flow",
}

_MORPH_COL_MAP = {
    10: "green_view",
    11: "blue_view",
    12: "built_ratio",
    13: "sky_view",
    14: "color_richness",
    15: "edge_density",
    16: "skyline_variance",
}

# 九人体感：每 7 列一人，从 Q(=17) 起
_PERSON_START_COLS = [17, 24, 31, 38, 45, 52, 59, 66, 73]


def image_key(name: str | Path) -> str:
    """取图片名称前 26 位（不含路径、保留 stem 前缀）。"""
    stem = Path(str(name)).name
    if Path(stem).suffix.lower() in {e.lower() for e in IMAGE_EXTENSIONS}:
        stem = Path(stem).stem
    return stem[:IMAGE_KEY_LEN]


def list_asset_images(assets_dir: Path | None = None) -> list[str]:
    """列出 assets 中可选图片文件名。"""
    root = Path(assets_dir or ASSETS_DIR)
    if not root.exists():
        return []
    files = [
        p.name
        for p in sorted(root.iterdir())
        if p.is_file() and p.suffix in IMAGE_EXTENSIONS
    ]
    return files


def list_scene_choices() -> list[str]:
    """
    下拉选项：优先 assets 文件名；若为空则用边缘密度图 stem（去 edge_ 前缀）
    以便在尚未放入原图时仍可联调指标与分析图。
    """
    assets = list_asset_images()
    if assets:
        return assets

    choices: list[str] = []
    if EDGE_MAPS_DIR.exists():
        for p in sorted(EDGE_MAPS_DIR.glob("edge_*")):
            # edge_XXX.png → XXX
            name = p.name
            if name.lower().startswith("edge_"):
                stem = Path(name[5:]).stem
                choices.append(stem)
    return choices


def _cell(row: tuple[Any, ...], col_1based: int) -> Any:
    idx = col_1based - 1
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _as_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _as_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _parse_person_name(header: str, index: int) -> tuple[str, str]:
    """从「体验感受指标（姓名…）」解析 (person_id, person_name)。"""
    text = (header or "").strip()
    m = re.search(r"[（(]([^）)]+)", text)
    raw = (m.group(1) if m else f"参与者{index}").strip()
    # 去掉性别等逗号后缀：段柔菲，女 → 段柔菲
    name = re.split(r"[,，]", raw, maxsplit=1)[0].strip() or f"参与者{index}"
    return f"p{index}", name


@lru_cache(maxsize=1)
def _load_workbook_rows(xlsx_path: str) -> tuple[tuple[Any, ...], ...]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        return tuple(tuple(row) for row in ws.iter_rows(values_only=True))
    finally:
        wb.close()


def clear_metrics_cache() -> None:
    _load_workbook_rows.cache_clear()


def load_metrics_row(image_name: str, xlsx_path: Path | None = None) -> dict[str, Any]:
    """
    按图片名前 26 位在 B 列找唯一行，返回情景 / 形态 / 九人体感。
    """
    path = Path(xlsx_path or FILLED_METRICS_XLSX)
    if not path.exists():
        raise FileNotFoundError(f"找不到指标表: {path}")

    key = image_key(image_name)
    rows = _load_workbook_rows(str(path.resolve()))
    if len(rows) < 3:
        raise ValueError(f"指标表行数不足: {path}")

    category_row = rows[0]
    header_row = rows[1]

    matches: list[tuple[int, tuple[Any, ...]]] = []
    for i, row in enumerate(rows[2:], start=3):
        b = _cell(row, 2)
        if b is None:
            continue
        if str(b).strip()[:IMAGE_KEY_LEN] == key:
            matches.append((i, row))

    if not matches:
        raise KeyError(f"在 filled_metrics.xlsx B 列未找到前缀 {key!r} 对应的行")
    if len(matches) > 1:
        line_nos = [m[0] for m in matches]
        raise KeyError(f"前缀 {key!r} 在 B 列匹配到多行: {line_nos}，请检查数据唯一性")

    excel_row_no, row = matches[0]

    scene_context = {
        field: _as_str(_cell(row, col)) for col, field in _SCENE_COL_MAP.items()
    }

    morph_metrics = {
        field: _as_float(_cell(row, col)) for col, field in _MORPH_COL_MAP.items()
    }

    persons: list[dict[str, Any]] = []
    for idx, start_col in enumerate(_PERSON_START_COLS, start=1):
        cat = _cell(category_row, start_col)
        person_id, person_name = _parse_person_name(_as_str(cat), idx)
        exp_vals = []
        for offset in range(7):
            exp_vals.append(_as_float(_cell(row, start_col + offset), default=3.0))
        # 裁剪到 1–5
        exp_vals = [min(5.0, max(1.0, v)) for v in exp_vals]
        experience = dict(zip(EXPERIENCE_KEYS, exp_vals))
        persons.append(
            {
                "person_id": person_id,
                "person_name": person_name,
                "experience": experience,
            }
        )

    return {
        "image_key": key,
        "excel_row": excel_row_no,
        "scene_id": _as_str(_cell(row, 1)),
        "excel_image_name": _as_str(_cell(row, 2)),
        "scene_context": scene_context,
        "morph_metrics": morph_metrics,
        "persons": persons,
        "headers": {
            "morph": {field: _as_str(_cell(header_row, col)) for col, field in _MORPH_COL_MAP.items()},
            "scene": {field: _as_str(_cell(header_row, col)) for col, field in _SCENE_COL_MAP.items()},
        },
    }


def _find_prefixed_map(directory: Path, prefix: str, stem_or_name: str) -> str | None:
    """在目录中找 prefix + 名称 对应的分析图，返回路径字符串。"""
    if not directory.exists():
        return None

    key = image_key(stem_or_name)
    full_stem = Path(str(stem_or_name)).name
    if Path(full_stem).suffix.lower() in {e.lower() for e in IMAGE_EXTENSIONS}:
        full_stem = Path(full_stem).stem
    # 若传入已是 edge_xxx，去掉前缀
    for pfx in ("edge_", "seg_", "skyline_"):
        if full_stem.startswith(pfx):
            full_stem = full_stem[len(pfx) :]

    candidates = [
        directory / f"{prefix}{full_stem}.png",
        directory / f"{prefix}{full_stem}.jpg",
        directory / f"{prefix}{key}.png",
    ]
    for c in candidates:
        if c.is_file():
            return str(c.resolve())

    globbed = sorted(directory.glob(f"{prefix}{key}*"))
    files = [p for p in globbed if p.is_file()]
    if len(files) == 1:
        return str(files[0].resolve())
    if len(files) > 1:
        # 优先完整 stem 包含关系更长的匹配
        exactish = [p for p in files if full_stem in p.stem]
        if len(exactish) == 1:
            return str(exactish[0].resolve())
        return str(files[0].resolve())
    return None


def resolve_asset_path(image_name: str, assets_dir: Path | None = None) -> str | None:
    """在 assets 中按完整文件名或前 26 位前缀解析原图路径。"""
    root = Path(assets_dir or ASSETS_DIR)
    if not root.exists():
        return None

    name = Path(str(image_name)).name
    exact = root / name
    if exact.is_file():
        return str(exact.resolve())

    # 无扩展名时尝试补全
    if not Path(name).suffix:
        for ext in (".jpg", ".jpeg", ".png", ".webp", ".JPG", ".PNG"):
            cand = root / f"{name}{ext}"
            if cand.is_file():
                return str(cand.resolve())

    key = image_key(name)
    hits = [
        p
        for p in root.iterdir()
        if p.is_file() and p.suffix in IMAGE_EXTENSIONS and p.name[:IMAGE_KEY_LEN] == key
    ]
    if len(hits) == 1:
        return str(hits[0].resolve())
    if len(hits) > 1:
        # 优先 stem 完全等于传入 stem
        stem = Path(name).stem if Path(name).suffix else name
        for p in hits:
            if p.stem == stem:
                return str(p.resolve())
        return str(sorted(hits)[0].resolve())
    return None


def resolve_scene_images(image_name: str) -> dict[str, str | None]:
    """解析原图 + 边缘密度 / 语义分割 / 天际线 三张分析图。"""
    asset = resolve_asset_path(image_name)
    # 用于匹配分析图的 stem：优先 assets 文件 stem，否则用选择名
    stem_ref = Path(asset).stem if asset else Path(str(image_name)).name
    if Path(stem_ref).suffix:
        stem_ref = Path(stem_ref).stem
    for pfx in ("edge_", "seg_", "skyline_"):
        if stem_ref.startswith(pfx):
            stem_ref = stem_ref[len(pfx) :]

    return {
        "original": asset,
        "edge_map": _find_prefixed_map(EDGE_MAPS_DIR, "edge_", stem_ref),
        "seg_map": _find_prefixed_map(SEG_MAPS_DIR, "seg_", stem_ref),
        "skyline_map": _find_prefixed_map(SKYLINE_MAPS_DIR, "skyline_", stem_ref),
    }


def average_experience(persons: list[dict[str, Any]]) -> dict[str, float]:
    """多人体验简单均值，用作前端目标滑块初值参考。"""
    if not persons:
        return {k: 3.0 for k in EXPERIENCE_KEYS}
    acc = {k: 0.0 for k in EXPERIENCE_KEYS}
    n = 0
    for p in persons:
        exp = p.get("experience") or {}
        n += 1
        for k in EXPERIENCE_KEYS:
            acc[k] += float(exp.get(k, 3.0))
    return {k: round(acc[k] / max(n, 1), 2) for k in EXPERIENCE_KEYS}


def load_scene_bundle(image_name: str) -> dict[str, Any]:
    """一次加载：指标行 + 图片路径，供前端 / API 使用。"""
    metrics = load_metrics_row(image_name)
    images = resolve_scene_images(image_name)
    return {
        **metrics,
        "images": images,
        "experience_average": average_experience(metrics["persons"]),
    }
